"""
将文件夹下的所有图片插入到 Excel 表格中，并嵌入到指定单元格中。
用户可以输入列宽和行高，未输入时使用默认值。
"""

import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

def set_column_dimensions(ws, target_column, column_width, row_height):
    """
    设置指定列的宽度和所有行的高度
    """
    # 设置列宽
    ws.column_dimensions[target_column].width = column_width
    # 设置所有行的高度
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = row_height

def get_cell_dimensions(ws, column, row):
    """
    获取单元格的宽度和高度（以像素为单位）
    """
    # 获取列宽和行高
    column_width = ws.column_dimensions[column].width or 10  # 默认宽度为 10
    row_height = ws.row_dimensions[row].height or 15  # 默认高度为 15

    # 将列宽和行高转换为像素
    # Excel 列宽和行高的单位与像素的换算公式
    cell_width = column_width * 7.5  # 1 列宽约等于 7.5 像素
    cell_height = row_height * 0.75  # 1 行高约等于 0.75 像素
    return int(cell_width), int(cell_height)

def resize_image(image_path, target_width, target_height):
    """
    调整图片大小以适应目标宽度和高度，同时保持图片比例
    """
    with PILImage.open(image_path) as img:
        # 获取图片的原始宽度和高度
        original_width, original_height = img.size

        # 计算缩放比例，保持图片比例
        scale = max(target_width / original_width, target_height / original_height)  # 使用 max 确保图片填满单元格
        new_width = int(original_width * scale)
        new_height = int(original_height * scale)

        # 调整图片大小
        resized_img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
        resized_path = f"{image_path}_resized.png"
        resized_img.save(resized_path, "PNG")
        return resized_path

def insert_images_to_excel(folder_path, target_column, column_width=15, row_height=84):
    """
    将文件夹下的图片插入到 Excel 表格中，并嵌入到指定单元格中
    """
    # 获取当前文件夹下的所有图片和 Excel 文件
    images = sorted([f for f in os.listdir(folder_path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))])
    excel_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.xlsx')]

    if not images:
        print("当前文件夹下没有找到图片文件！")
        return

    if not excel_files:
        print("当前文件夹下没有找到 Excel 文件！")
        return

    # 假设只有一个 Excel 文件
    excel_file = excel_files[0]
    wb = load_workbook(os.path.join(folder_path, excel_file))
    ws = wb.active

    # 设置列宽和行高
    set_column_dimensions(ws, target_column, column_width, row_height)

    # 记录临时生成的图片路径
    temp_images = []

    # 遍历图片文件，将图片插入到对应的行
    for image_file in images:
        # 提取图片编号，例如 01.JPG -> 1
        image_number = int(os.path.splitext(image_file)[0])
        # 查找 Excel 中对应编号的行
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == image_number:
                # 获取单元格的宽度和高度
                cell_width, cell_height = get_cell_dimensions(ws, target_column, row[0].row)

                # 调整图片大小并插入到目标列
                resized_image_path = resize_image(os.path.join(folder_path, image_file), cell_width, cell_height)
                temp_images.append(resized_image_path)  # 记录临时图片路径
                img = Image(resized_image_path)
                img.anchor = f"{target_column}{row[0].row}"  # 设置图片的锚点为目标单元格
                ws.add_image(img)
                print(f"已将图片 {image_file} 嵌入到单元格 {target_column}{row[0].row}")
                break

    # 保存 Excel 文件
    wb.save(os.path.join(folder_path, excel_file))
    print(f"图片已成功嵌入到 {excel_file} 中！")

    # 删除临时生成的图片
    for temp_image in temp_images:
        try:
            os.remove(temp_image)
            print(f"已删除临时图片: {temp_image}")
        except Exception as e:
            print(f"删除临时图片 {temp_image} 时发生错误: {e}")

if __name__ == "__main__":
    print("****************************\n使用说明：\n  1.将所有需要插入的图片和excel放在当前文件夹下。\n  2.输入要插入图片的列（例如:E），以及列宽和行高（可选，默认直接回车即可）\n注意:1.excel文件格式必须为xlsx。2.图片文件格式必须为jpg、jpeg或png,且编号正确。\n****************************")

    folder_path = os.getcwd()  # 当前文件夹路径
    target_column = input("请输入要插入图片的列（例如:E）：").strip().upper()
    column_width = input("请输入列宽（默认 15）：").strip()
    row_height = input("请输入行高（默认 84）：").strip()

    # 转换输入值为数字，若未输入则使用默认值
    column_width = float(column_width) if column_width else 15
    row_height = float(row_height) if row_height else 84

    insert_images_to_excel(folder_path, target_column, column_width, row_height)